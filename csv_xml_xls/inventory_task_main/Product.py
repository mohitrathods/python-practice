import os
import xlsxwriter
import xlrd
import xlwt
import openpyxl
from openpyxl import load_workbook
from xlutils.copy import copy

class Product:

    def add_extra_cols(self):
        wb = xlrd.open_workbook("/home/setu/PycharmProjects/python-practice/csv_xml_xls/inventory_task_main/data/sales.xls")
        ws = wb.sheet_by_index(0)
        copyewb = copy(wb)
        copy_sh = copyewb.get_sheet(0)
        copy_sh.write(0,5,'CGST')
        copy_sh.write(0,6,'SGST')
        copy_sh.write(0,7,'IGST')
        copy_sh.write(0,8,'Total Tax')
        copyewb.save("/home/setu/PycharmProjects/python-practice/csv_xml_xls/inventory_task_main/data/sales.xls")

    # -------------------------------------------------------------------------------------
    # create xlsx database/file with fields
    def createXlDatabase(self):
        # fields only needed and call this fun only if xl file is not available at location
        path = '/home/setu/PycharmProjects/python-practice/csv_xml_xls/inventory_task_main/data/'
        files = {
            'master_inventory.xls': ['product_name', 'product_sku', 'quantity', 'purchase_price', 'total_amount',
                                      'vendor_name', 'sales_price'],
            'purchase.xls': ['product_name', 'product_sku', 'quantity', 'purchase_price', 'total_amount','CGST','SGST','IGST', 'total_tax'],
            'inventory_movement.xls': ['product_name', 'date_of_changes', 'quantity_changes', 'price', 'customer_name',
                                        'vendor', 'purchase_or_cell'],
            'sales.xls': ['product_name', 'product_sku', 'quantity', 'purchase_price', 'total_amount','CGST','SGST','IGST', 'total_tax']
        }
        for file_name, values in files.items():
            if os.path.exists(path + file_name):
                if os.path.isfile(path + file_name):
                    pass
            else:
                # workbook = xlsxwriter.Workbook(path + file_name)
                # worksheet = workbook.add_worksheet("Sheet1")
                #WITH XLWT
                wb = xlwt.Workbook(encoding='utf8')
                sh = wb.add_sheet("Sheet1")
                c = 0
                for each in values:
                    sh.write(0, c, each)
                    c += 1
                wb.save(path+file_name)


    # -------------------------------------------------------------------------------------

    # get product input from user add this data to main inventory
    def getInput(self):
        # self.createXlDatabase()
        li=[]

        workbook = xlrd.open_workbook("/home/setu/PycharmProjects/python-practice/csv_xml_xls/inventory_task_main/data/master_inventory.xls")
        worksheet = workbook.sheet_by_index(0)
        nth_row = worksheet.nrows
        nth_col = worksheet.ncols

        copy_wb = copy(workbook)
        copy_sheet = copy_wb.get_sheet(0)

        sku = str(input("Enter sku : "))

        flag = False
        j = 0
        x = None
        for i in range(nth_row):
            eachrow = worksheet.row(i)
            if eachrow[1].value == sku:
                flag = True
                j = i
                x = eachrow
                break

        if flag:
            # print(x)
            li.clear()
            print("UPDATE data here")
            add_qty = float(input(f"You have {x[2].value} quantity in stock > Add more stock > add quantity : "))
            total_qty = x[2].value + add_qty
            final_total = total_qty * x[3].value
            copy_sheet.write(j, 2, total_qty) # row,col,value
            copy_sheet.write(j, 4, final_total) # row,col,value
            copy_wb.save("/home/setu/PycharmProjects/python-practice/csv_xml_xls/inventory_task_main/data/master_inventory.xls")
            # what is to be added, append in li and return li
            # pname, sku, qty(current added), purchase_price(get from master_inventory), total(qty*purchaseprice)
            li.append(x[0].value)
            li.append(x[1].value)
            li.append(add_qty)
            li.append(x[3].value)
            li.append(add_qty*x[3].value)
            li.append(x[5].value)
            # print(li)

        if not flag:
            li.clear()
            print("ADD NEW DATA")
            product_name = str(input("Enter product name : "))
            quantity = int(input("Enter quantity : "))
            vendor_name = str(input("Enter vendor name : "))
            purchase_price = float(input("Enter purchase price : "))
            total_amount = purchase_price * quantity
            sell_price = float(input("Enter selling price : "))
            arr = [product_name, sku, quantity, purchase_price, total_amount, vendor_name, sell_price]
            c = 0
            for each in arr:
                copy_sheet.write(nth_row,c,each)
                li.append(each)
                c+=1
            copy_wb.save("/home/setu/PycharmProjects/python-practice/csv_xml_xls/inventory_task_main/data/master_inventory.xls")
        # op = li.copy()
        # li.clear()
        return li

    # -------------------------------------------------------------------------------------

# prd = Product()
# prd.createXlDatabase()