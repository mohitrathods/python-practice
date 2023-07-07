import os
import xlsxwriter
import xlrd
import xlwt
import openpyxl
from openpyxl import load_workbook
from xlutils.copy import copy

def createXlDatabase(self):
    # fields only needed and call this fun only if xl file is not available at location
    path = '/home/setu/PycharmProjects/python-practice/csv_xml_xls/inventory_task_main/data/'
    files = {
        'master_inventory.xls': ['product_name', 'product_sku', 'quantity', 'purchase_price', 'total_amount',
                                 'vendor_name', 'sales_price'],
        'purchase.xls': ['product_name', 'product_sku', 'quantity', 'purchase_price', 'total_amount', 'CGST', 'SGST',
                         'IGST', 'total_tax'],
        'inventory_movement.xls': ['product_name', 'date_of_changes', 'quantity_changes', 'price', 'customer_name',
                                   'vendor', 'purchase_or_cell'],
        'sales.xls': ['product_name', 'product_sku', 'quantity', 'purchase_price', 'total_amount', 'CGST', 'SGST',
                      'IGST', 'total_tax']
    }
    for file_name, values in files.items():
        if os.path.exists(path + file_name):
            if os.path.isfile(path + file_name):
                pass
        else:
            # workbook = xlsxwriter.Workbook(path + file_name)
            # worksheet = workbook.add_worksheet("Sheet1")
            # WITH XLWT
            wb = xlwt.Workbook(encoding='utf8')
            sh = wb.add_sheet("Sheet1")
            c = 0
            for each in values:
                sh.write(0, c, each)
                c += 1
            wb.save(path + file_name)

def testing(self):
    #READ
    wb = xlrd.open_workbook("/home/setu/PycharmProjects/python-practice/csv_xml_xls/inventory_task_main/testing/inventory_movement.xls")
    ws = wb.sheet_by_name('Sheet1') #w1

    nth_row = ws.nrows
    nth_col = ws.ncols

    #WRITE
    copy_wb = copy(wb) # make writable copy of opened excel file
    copy_sh = copy_wb.get_sheet(0) # read the first sheet to withe within writable copy

def excel_func(self,workbook):
    workbook = xlrd.open_workbook(workbook)  # open each workbook
    worksheet = workbook.sheet_by_index(0)  # get sheet from sheet no
    nth_row = worksheet.nrows  # get nth row
    nth_col = worksheet.ncols  # get nth col
    copy_workbook = copy(workbook)  # copy each workbook : FOR SAVE OPERATION
    copy_worksheet = copy_workbook.get_sheet(0)  # copy each worksheet : FOR WRITING OPERATION
    l = [workbook, worksheet, nth_row, nth_col, copy_workbook, copy_worksheet]

    return l